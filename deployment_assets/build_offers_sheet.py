"""Generate the bulk-upload sheet for the launch offer catalogue.

Column headers must match the parser in public/admin/manage-offers.html exactly -
processBulkData() reads item['Merchant'], item['Bank Name'], etc. by name.

Sheet 1 "Offers"      - single-tier offers, ready for Bulk Upload.
Sheet 2 "Multi-Tier"  - offers whose terms differ per product. Bulk upload cannot
                        express tiers, so these are entered via the manual form's
                        Offer Breakdown rows. Listed here so nothing is lost.
Sheet 3 "README"      - column meanings and warnings.

Every row was researched from the issuer's own public offer page; the Offer Source
column holds that URL so each can be re-verified before publishing.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "Merchant", "Bank Name", "Category", "Payment Type",
    "Discount Type", "Discount Value", "Max Discount", "Min Order",
    "Description", "Terms & C", "Coupon Code", "Code On Site",
    "Merchant URL", "Offer Source", "Valid Until",
]

SBI = "https://www.sbicard.com/en/personal/offer/"
ICICI = "https://www.icici.bank.in/offers/"
AU = "https://offers.au.bank.in/offers/offer-details/"


def row(merchant, bank, category, payment, dtype, dvalue, maxd, mino,
        desc, code, on_site, murl, source, valid):
    return {
        "Merchant": merchant, "Bank Name": bank, "Category": category,
        "Payment Type": payment, "Discount Type": dtype, "Discount Value": dvalue,
        "Max Discount": maxd, "Min Order": mino, "Description": desc,
        "Terms & C": "", "Coupon Code": code, "Code On Site": on_site,
        "Merchant URL": murl, "Offer Source": source, "Valid Until": valid,
    }


ROWS = [
    # ---------- SBI Card ----------
    row("BigBasket", "SBI Card", "Groceries", "Credit Card", "Percentage", 10, 250, 999,
        "10% instant discount on groceries at BigBasket. Valid 26th to month-end, once per card per month.",
        "SBICC", "", "https://www.bigbasket.com", SBI + "big-basket-29july26.page", "2026-09-30"),
    row("Apple", "SBI Card", "Shopping", "Credit Card", "Flat", 15000, "", "",
        "Up to Rs 15,000 instant discount on iPhone, MacBook, iPad, Watch and Pods. EMI transactions only.",
        "", "", "https://www.apple.com/in/", SBI + "apple-9july26.page", "2026-09-26"),
    row("IGP", "SBI Card", "Shopping", "Credit Card", "Percentage", 20, "", 1199,
        "20% instant discount on gifts and flowers at IGP.",
        "SBIIGP20", "", "https://www.igp.com", SBI + "igp-1june26.page", "2027-03-31"),
    row("ClearTax", "SBI Card", "Shopping", "Credit Card", "Percentage", 75, "", "",
        "Flat 75% off ClearTax income tax filing plans, self-filing and expert-assisted.",
        "", "Yes", "https://cleartax.in", SBI + "clear-tax-4july26.page", "2027-03-31"),
    row("MediBuddy", "SBI Card", "Shopping", "Credit Card", "Flat", 3600, "", "",
        "3 months unlimited online doctor consultations for 6 family members at Rs 1,399 instead of Rs 4,999.",
        "SBICMBG", "", "https://www.medibuddy.in", SBI + "medibuddy-16july26.page", "2027-03-31"),
    row("MediBuddy", "SBI Card", "Shopping", "Credit Card", "Flat", 600, "", "",
        "Full body health checkup with 85 parameters at Rs 1,899 instead of Rs 2,499.",
        "SBIC600", "", "https://www.medibuddy.in", SBI + "medibuddy-16july26.page", "2027-03-31"),
    row("Flipkart", "SBI Card", "Shopping", "Credit Card", "Percentage", 10, 1000, 4990,
        "10% instant discount on Flipkart. EMI transactions only. Cap Rs 750 on mobiles, Rs 1,000 other categories.",
        "", "", "https://www.flipkart.com", SBI + "flipkart-28july26.page", "2026-07-31"),
    row("Amazon", "SBI Card", "Shopping", "Credit Card", "Percentage", 10, 1000, 5000,
        "10% instant discount on Amazon. EMI transactions only. Cap Rs 1,000 per transaction, Rs 4,000 per card.",
        "", "", "https://www.amazon.in", SBI + "amazon-15july26.page", "2026-07-31"),

    # ---------- ICICI Bank ----------
    row("TaxBuddy", "ICICI Bank", "Shopping", "Credit Card, Debit Card", "Percentage", 20, "", "",
        "20% off ITR filing, tax planning, notice management and GST services. Also valid on Net Banking.",
        "", "Yes", "https://www.taxbuddy.com", ICICI + "tax-buddy-discount-offer", "2027-04-30"),
    row("Surat Diamond", "ICICI Bank", "Shopping", "Credit Card, Debit Card", "Percentage", 20, "", 2000,
        "20% instant discount on jewellery at Surat Diamond. Excludes bullion, coins and solitaires.",
        "ICICIBANK20", "", "https://www.suratdiamond.com",
        ICICI + "get-20-percent-instant-discount-at-surat-diamond", "2027-03-31"),
    row("Carzonrent", "ICICI Bank", "Travel", "Credit Card, Debit Card", "Percentage", 30, "", "",
        "Flat 30% off all Carzonrent rides. First eligible transaction per card per month. Not valid on EMI.",
        "", "", "https://www.carzonrent.com", ICICI + "offer-on-carzonrent", "2026-09-30"),
    row("Insight Vacations", "ICICI Bank", "Travel", "Credit Card, Debit Card", "Percentage", 10, 25000, "",
        "10% off Insight Vacations tours, up to Rs 25,000 per person. Offline bookings only, minimum two adults.",
        "", "", "https://www.insightvacations.com", ICICI + "insigh-vacation", "2027-09-30"),

    # ---------- AU Bank ----------
    row("Swiggy", "AU Bank", "Dining", "Credit Card", "Percentage", 6, 60, 300,
        "6% off food orders on Swiggy. Once per card per week. Not valid on AU co-branded or corporate cards.",
        "AUCC60", "", "https://www.swiggy.com",
        AU + "5f4f7db2-24fd-4904-bec7-0820c17de756/generic/swiggy", "2026-09-30"),
    row("Zomato", "AU Bank", "Dining", "Credit Card", "Percentage", 3, "", 199,
        "Flat 3% off food orders on Zomato. Can be availed multiple times a month.",
        "AUCC3", "", "https://www.zomato.com",
        AU + "dc8b79db-838b-447b-8ac9-d5d36a264595/generic/zomato", "2026-09-30"),
    row("Behrouz Biryani", "AU Bank", "Dining", "Credit Card, Debit Card", "Percentage", 30, 299, 299,
        "Up to 30% off on Behrouz Biryani via the EatSure app.",
        "", "Yes", "https://www.behrouzbiryani.com",
        AU + "b7a2623d-c7b5-4336-a33c-21c9c5d10500/generic/behrouz-biryani", "2026-12-31"),
    row("PVR INOX", "AU Bank", "Entertainment", "Credit Card", "Flat", 125, "", 500,
        "Flat Rs 125 off movie tickets at PVR INOX. Once per card per month, movie tickets only.",
        "", "Yes", "https://www.pvrcinemas.com",
        AU + "1bf349b7-46dd-4dd4-9d41-fec5a404544c/generic/pvr-inox", "2026-09-30"),
    row("BookMyShow", "AU Bank", "Entertainment", "Debit Card", "Flat", 500, 500, "",
        "Buy one get one free on movie and non-movie tickets with AU ivy Debit Card. Minimum two tickets. "
        "6 times per month on movies, twice on non-movies.",
        "", "Yes", "https://in.bookmyshow.com",
        AU + "e642a7d3-8ebf-472c-9ccb-5aea45519fb9/generic/bookmyshow-", "2026-12-31"),
    row("BigBasket", "AU Bank", "Groceries", "Debit Card", "Percentage", 15, 300, 1500,
        "15% off on BigBasket with AU ivy and 'M' Circle Debit Cards. Valid 1st to 25th of each month, "
        "once per card during the offer period.",
        "AUDC15", "", "https://www.bigbasket.com",
        AU + "02dcec10-5bd2-480c-a284-9d4e79f2ff61/generic/bigbasket", "2026-12-31"),
    row("Netmeds", "AU Bank", "Shopping", "Credit Card, Debit Card", "Percentage", 8, "", "",
        "Additional 8% off medicines at Netmeds, on top of the standard 12% platform discount.",
        "", "Yes", "https://www.netmeds.com",
        AU + "74700860-ef2d-46d9-8036-cec4d84bb0e0/generic/netmeds", "2026-12-31"),
]

# Offers whose terms differ per product - bulk upload has no tier support, so these
# are documented for manual entry rather than silently flattened to one number.
TIER_HEADERS = ["Merchant", "Bank Name", "Category", "Payment Type", "Headline",
                "Tier", "Tier Discount", "Tier Cap", "Tier Min Order", "Coupon Code",
                "Key Restriction", "Merchant URL", "Offer Source", "Valid Until"]

TIERED = [
    ("MakeMyTrip", "AU Bank", "Travel", "Credit Card", "Up to 12% off",
     [("Domestic Flight", "12%", 1000, 5000), ("Domestic Hotel", "10%", 1000, 4000)],
     "MMTAU", "Valid on bookings made on TUESDAYS only. Once per card per month.",
     "https://www.makemytrip.com", AU + "4b935cd6-a71e-4bff-8faf-c31a3ea78407/generic/makemytrip", "2026-09-30"),
    ("EaseMyTrip", "AU Bank", "Travel", "Credit Card", "Up to 12% off",
     [("Domestic Flight", "12%", 1000, 5000), ("International Flight", "10%", 4500, 20000),
      ("Domestic Hotel", "10%", 1000, 4000), ("International Hotel", "10%", 3000, 15000)],
     "", "Valid on bookings made on THURSDAYS only. Once per card per month.",
     "https://www.easemytrip.com", AU + "42500c6e-ef10-4b36-b537-79ffb2f3443b/generic/easemytrip", "2026-09-30"),
    ("Ixigo", "AU Bank", "Travel", "Credit Card", "Up to 12% off",
     [("Domestic Flight", "12%", 1000, 5000), ("International Flight", "10%", 4500, 20000),
      ("Domestic Hotel", "10%", 1000, 4000), ("International Hotel", "10%", 3000, 15000)],
     "", "Valid on bookings made on SUNDAYS only. Once per card per month.",
     "https://www.ixigo.com", AU + "16614916-ba63-41f8-be95-42f4200e9fac/generic/ixigo", "2026-09-30"),
    ("Yatra", "AU Bank", "Travel", "Credit Card", "Up to Rs 10,000 off",
     [("Domestic Flight - 1 ticket", "Flat Rs 799", "", ""),
      ("Domestic Flight - 2 tickets", "Flat Rs 1,598", "", ""),
      ("Domestic Flight - 3 tickets", "Flat Rs 2,397", "", ""),
      ("International Flight - under Rs 30,000", "Flat Rs 2,000", "", ""),
      ("International Flight - Rs 30,000+", "8%", 10000, 30000)],
     "", "Once per card per month. Not valid on cards starting 406977503 or 653062913.",
     "https://www.yatra.com", AU + "0a577abd-7dc0-4d48-a626-00b4376f8895/generic/yatra", "2026-09-30"),
    ("Zepto", "AU Bank", "Groceries", "Credit Card", "Up to Rs 1,000 off",
     [("Grocery", "Flat Rs 50", "", 500), ("Electronics", "5%", 1000, 5000)],
     "ZEPAUCC (grocery) / ZEPAUHV (electronics)", "Once per card per month.",
     "https://www.zeptonow.com", AU + "bb0b8fda-486b-4821-bf66-361214f4d8f7/generic/zepto", "2026-09-30"),
]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
BODY_FONT = Font(name="Arial")


def style_sheet(ws, headers, widths):
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 18)
    ws.freeze_panes = "A2"


wb = Workbook()

# --- Sheet 1: bulk-uploadable offers ---
ws = wb.active
ws.title = "Offers"
ws.append(HEADERS)
for r in ROWS:
    ws.append([r.get(h, "") for h in HEADERS])
style_sheet(ws, HEADERS, {
    "Merchant": 18, "Bank Name": 13, "Category": 14, "Payment Type": 22,
    "Discount Type": 13, "Discount Value": 13, "Max Discount": 12, "Min Order": 10,
    "Description": 64, "Terms & C": 16, "Coupon Code": 14, "Code On Site": 12,
    "Merchant URL": 30, "Offer Source": 46, "Valid Until": 12})

# --- Sheet 2: multi-tier offers for manual entry ---
mt = wb.create_sheet("Multi-Tier")
mt.append(TIER_HEADERS)
for (merch, bank, cat, pay, headline, tiers, code, restriction, murl, src, valid) in TIERED:
    for idx, (label, disc, cap, mino) in enumerate(tiers):
        mt.append([
            merch if idx == 0 else "", bank if idx == 0 else "", cat if idx == 0 else "",
            pay if idx == 0 else "", headline if idx == 0 else "",
            label, disc, cap, mino,
            code if idx == 0 else "", restriction if idx == 0 else "",
            murl if idx == 0 else "", src if idx == 0 else "", valid if idx == 0 else "",
        ])
style_sheet(mt, TIER_HEADERS, {
    "Merchant": 16, "Bank Name": 11, "Category": 12, "Payment Type": 16, "Headline": 20,
    "Tier": 32, "Tier Discount": 14, "Tier Cap": 10, "Tier Min Order": 13,
    "Coupon Code": 30, "Key Restriction": 56, "Merchant URL": 28,
    "Offer Source": 44, "Valid Until": 12})

# --- Sheet 3: README ---
notes = wb.create_sheet("README")
notes["A1"] = "OfferLens offer catalogue - how to use"
notes["A1"].font = Font(name="Arial", bold=True, size=14)
guidance = [
    "",
    "SHEET 'Offers' - upload these via admin panel > Bulk Upload. Only the first sheet is read.",
    "SHEET 'Multi-Tier' - enter these by hand; bulk upload cannot express per-product tiers.",
    "   Use the manual form's 'Offer Breakdown' rows, one per tier.",
    "",
    "BEFORE UPLOADING: verify each row against its Offer Source URL. These were researched",
    "automatically from issuer portals and terms change without notice.",
    "",
    "Column notes:",
    "  Bank Name      Must match the Smart Wallet list exactly, or card filtering won't work.",
    "  Category       Must be one of the app's 7 categories, or the offer appears under no tab.",
    "  Discount Type  Percentage / Flat / Cashback.",
    "  Max Discount   The cap for 'X% off up to Rs Y'. Blank if uncapped.",
    "  Min Order      Blank rather than 0 when there is no minimum.",
    "  Terms & C      Left blank deliberately. After upload, open each offer with Edit and use",
    "                 'Generate from offer details' - do not paste the issuer's own wording.",
    "  Code On Site   'Yes' when a code is required but only shown on the issuer's site.",
    "  Offer Source   The issuer's page. Merchant URL is the merchant's own site.",
    "  Valid Until    YYYY-MM-DD.",
    "",
    "WATCH OUT:",
    "  - BookMyShow and BigBasket (AU) are DEBIT card offers (ivy / 'M' Circle), not credit.",
    "  - Travel offers have day-of-week limits: MakeMyTrip Tue, EaseMyTrip Thu, Ixigo Sun.",
    "  - Flipkart and Amazon (SBI) are EMI-only and expire 31 Jul 2026 - drop if launching later.",
    "  - AU offers exclude AU co-branded and corporate cards.",
    "  - Most offers are once per card per month.",
]
for i, line in enumerate(guidance, start=2):
    notes[f"A{i}"] = line
    notes[f"A{i}"].font = Font(name="Arial", bold=line.endswith(":") or line.startswith("SHEET"))
notes.column_dimensions["A"].width = 110

out = "deployment_assets/offerlens_bulk_offers.xlsx"
wb.save(out)
print(f"wrote {out}: {len(ROWS)} bulk rows, {len(TIERED)} multi-tier offers")
