"""
Builds the Batch 2 bulk-upload workbook for OfferLens.

Every row here was read from the issuer's own HTTPS site, with each `.bank.in`
domain confirmed by 301 redirect from the bank's long-established primary domain
rather than from a search ranking. Source URLs are carried in the sheet so each
row can be re-checked against its origin before publishing.

Nothing in this file is inferred. Where an issuer published no cap, no minimum,
or no end date, the cell is left blank rather than filled with a guess.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

COLUMNS = [
    "Merchant", "Bank Name", "Category", "Payment Type", "Discount Type",
    "Discount Value", "Max Discount", "Min Order", "Description", "Terms & C",
    "Coupon Code", "Code On Site", "Merchant URL", "Offer Source", "Valid Until",
]

RBL = "https://www.rbl.bank.in/offers"
BOB = "https://www.bobcard.co.in/offers"
KOTAK = "https://www.kotak.bank.in/en/offers.html"
HSBC = "https://www.hsbc.co.in/credit-cards/offers/"
SC = "https://www.sc.bank.in/credit-cards/digismart-card/"
AMEX = "https://www.americanexpress.com/en-in/benefits/emi/emi-offers/in-store/index.html"
ONECARD = "https://www.getonecard.app/sbmoffers/"

# Merchant, Bank, Category, PaymentType, DiscountType, Value, MaxDisc, MinOrder,
# Description, Terms, Code, CodeOnSite, MerchantURL, Source, ValidUntil
OFFERS = [
    # --- RBL Bank ---------------------------------------------------------
    ("Oven Story", "RBL Bank", "Food & Dining", "Credit Card", "Percentage", 50, 125, 249,
     "50% off on Oven Story orders with RBL Bank cards.",
     "Capped at Rs 125 per transaction. Minimum order Rs 249.",
     "", "TRUE", "https://www.ovenstory.in", RBL, "2026-09-30"),
    ("The Good Bowl", "RBL Bank", "Food & Dining", "Credit Card", "Percentage", 40, 110, 199,
     "40% off on The Good Bowl orders with RBL Bank cards.",
     "Capped at Rs 110 per transaction. Minimum order Rs 199.",
     "", "TRUE", "https://www.thegoodbowl.in", RBL, "2026-09-30"),
    ("Faasos", "RBL Bank", "Food & Dining", "Credit Card", "Percentage", 40, 110, 199,
     "40% off on Faasos orders with RBL Bank cards.",
     "Capped at Rs 110 per transaction. Minimum order Rs 199.",
     "", "TRUE", "https://www.faasos.com", RBL, "2026-09-30"),
    ("Behrouz Biryani", "RBL Bank", "Food & Dining", "Credit Card", "Percentage", 30, 125, 249,
     "30% off on Behrouz Biryani orders with RBL Bank cards.",
     "Capped at Rs 125 per transaction. Minimum order Rs 249.",
     "", "TRUE", "https://www.behrouzbiryani.com", RBL, "2026-09-30"),
    ("EatSure", "RBL Bank", "Food & Dining", "Credit Card", "Percentage", 30, "", 249,
     "30% off on EatSure orders with RBL Bank cards.",
     "Minimum order Rs 249. No maximum cap published by the issuer.",
     "", "TRUE", "https://www.eatsure.com", RBL, "2026-09-30"),
    ("Arata", "RBL Bank", "Health & Beauty", "Credit Card", "Percentage", 30, "", "",
     "30% off on Arata with RBL Bank cards.",
     "No cap or minimum order published by the issuer.",
     "", "TRUE", "https://www.arata.in", RBL, "2026-09-30"),

    # --- Bank of Baroda (BOBCARD) ----------------------------------------
    ("Cleartrip", "Bank of Baroda", "Travel", "EMI", "Percentage", 25, "", "",
     "25% instant discount on Cleartrip bookings with BOBCARD on EMI.",
     "EMI transactions only. Cap and minimum order not published by the issuer.",
     "BOBEMI", "FALSE", "https://www.cleartrip.com", BOB, "2026-09-27"),
    ("Flipkart Travel", "Bank of Baroda", "Travel", "EMI", "Percentage", 20, "", "",
     "20% instant discount on Flipkart Travel bookings with BOBCARD on EMI.",
     "EMI transactions only. Use BOBEMIDOM for domestic, BOBEMIINT for international.",
     "BOBEMIDOM", "FALSE", "https://www.flipkart.com/travel", BOB, "2026-08-30"),
    ("MakeMyTrip", "Bank of Baroda", "Travel", "EMI", "Percentage", 15, "", "",
     "15% instant discount on MakeMyTrip bookings with BOBCARD on EMI.",
     "EMI transactions only. Use MMTBOBEMI for domestic, MMTBOBINTEMI for international.",
     "MMTBOBEMI", "FALSE", "https://www.makemytrip.com", BOB, "2026-09-30"),
    ("Goibibo", "Bank of Baroda", "Travel", "EMI", "Percentage", 15, "", "",
     "15% instant discount on Goibibo bookings with BOBCARD on EMI.",
     "EMI transactions only. Use GOBOBEMI for domestic, GOBOBINTEMI for international.",
     "GOBOBEMI", "FALSE", "https://www.goibibo.com", BOB, "2026-09-30"),
    ("EaseMyTrip", "Bank of Baroda", "Travel", "EMI", "Percentage", 15, "", "",
     "15% instant discount on EaseMyTrip bookings with BOBCARD on EMI.",
     "EMI transactions only. Cap and minimum order not published by the issuer.",
     "BOBEMI", "FALSE", "https://www.easemytrip.com", BOB, "2026-09-30"),
    ("Ixigo", "Bank of Baroda", "Travel", "EMI", "Percentage", 12, "", "",
     "12% instant discount on Ixigo bookings with BOBCARD on EMI.",
     "EMI transactions only. Cap and minimum order not published by the issuer.",
     "IXIBOBEMI", "FALSE", "https://www.ixigo.com", BOB, "2026-09-30"),
    ("Paytm Travel", "Bank of Baroda", "Travel", "EMI", "Percentage", 12, "", "",
     "12% instant discount on Paytm flight bookings with BOBCARD on EMI.",
     "EMI transactions only. Use FLYBOB for domestic, INTLFLYBOB for international.",
     "FLYBOB", "FALSE", "https://www.paytm.com/flights", BOB, "2026-09-30"),
    ("Flipkart", "Bank of Baroda", "Shopping", "Credit Card", "Percentage", 10, "", "",
     "10% instant discount on Flipkart with BOBCARD.",
     "Cap and minimum order not published by the issuer.",
     "", "TRUE", "https://www.flipkart.com", BOB, "2026-08-31"),
    ("Amazon", "Bank of Baroda", "Shopping", "Credit Card", "Percentage", 10, "", "",
     "10% instant discount on Amazon with BOBCARD.",
     "Cap and minimum order not published by the issuer.",
     "", "TRUE", "https://www.amazon.in", BOB, "2026-08-31"),
    ("Surat Diamond", "Bank of Baroda", "Shopping", "Credit Card", "Flat", 1000, "", "",
     "Flat Rs 1,000 off at Surat Diamond with BOBCARD.",
     "Minimum order not published by the issuer.",
     "BOBCARD1000", "FALSE", "https://www.suratdiamond.com", BOB, "2027-03-29"),
    ("udChalo", "Bank of Baroda", "Travel", "Credit Card", "Percentage", 5, "", "",
     "5% discount on udChalo bookings with BOBCARD.",
     "Cap and minimum order not published by the issuer.",
     "UCBOB", "FALSE", "https://www.udchalo.com", BOB, "2026-09-30"),

    # --- Kotak Mahindra ---------------------------------------------------
    ("IFB", "Kotak Mahindra", "Electronics", "EMI", "Percentage", 20, 9000, "",
     "20% cashback on IFB appliances with Kotak Credit Card EMI.",
     "Credit Card EMI transactions only. Capped at Rs 9,000.",
     "", "TRUE", "https://www.ifbappliances.com", KOTAK, "2026-09-30"),
    ("Oppo", "Kotak Mahindra", "Electronics", "EMI", "Percentage", 10, 3000, "",
     "10% cashback on Oppo devices with Kotak Credit Card EMI.",
     "Credit Card EMI transactions only. Capped at Rs 3,000.",
     "", "TRUE", "https://www.oppo.com/in", KOTAK, "2026-09-30"),

    # --- HSBC -------------------------------------------------------------
    ("Zepto", "HSBC", "Grocery", "Credit Card", "Flat", 100, "", 999,
     "Flat Rs 100 off on Zepto with HSBC Credit Cards.",
     "Minimum order Rs 999.",
     "", "TRUE", "https://www.zeptonow.com", HSBC, "2026-12-31"),
    ("Lakme Salon", "HSBC", "Health & Beauty", "Credit Card", "Flat", 1100, "", 3000,
     "Flat Rs 1,100 off at Lakme Salon with HSBC Credit Cards.",
     "Minimum order Rs 3,000.",
     "", "TRUE", "https://www.lakmesalon.in", HSBC, "2026-12-31"),
    ("Emirates", "HSBC", "Travel", "Credit Card", "Flat", 12000, "", "",
     "Up to Rs 12,000 off on Emirates bookings with HSBC Credit Cards.",
     "Rs 12,000 is the maximum benefit; actual discount varies by fare class.",
     "INHSBC1", "FALSE", "https://www.emirates.com/in", HSBC, "2026-08-31"),

    # --- Standard Chartered (DigiSmart) -----------------------------------
    ("Blinkit", "Standard Chartered", "Grocery", "Credit Card", "Percentage", 10, 1000, "",
     "10% off on Blinkit with the Standard Chartered DigiSmart Credit Card.",
     "DigiSmart Credit Card only. Capped at Rs 1,000 per month across 5 transactions.",
     "DIGISMART", "FALSE", "https://www.blinkit.com", SC, ""),
    ("Zomato", "Standard Chartered", "Food & Dining", "Credit Card", "Percentage", 10, 150, "",
     "10% off on Zomato with the Standard Chartered DigiSmart Credit Card.",
     "DigiSmart Credit Card only. Capped at Rs 150 per transaction.",
     "DIGISMART", "FALSE", "https://www.zomato.com", SC, ""),
    ("Yatra", "Standard Chartered", "Travel", "Credit Card", "Percentage", 20, 750, "",
     "20% off on Yatra domestic flights with the Standard Chartered DigiSmart Credit Card.",
     "DigiSmart Credit Card only. Domestic flights. Capped at Rs 750.",
     "DIGISMART", "FALSE", "https://www.yatra.com", SC, ""),

    # --- American Express -------------------------------------------------
    ("Oppo", "American Express", "Electronics", "EMI", "Flat", 17000, "", "",
     "Up to Rs 17,000 instant discount on Oppo with American Express EMI.",
     "In-store EMI transactions. Rs 17,000 is a ceiling, not a guaranteed amount.",
     "", "TRUE", "https://www.oppo.com/in", AMEX, "2026-09-30"),
    ("Asus", "American Express", "Electronics", "EMI", "Flat", 3000, "", "",
     "Up to Rs 3,000 instant discount on Asus with American Express EMI.",
     "In-store EMI over 3/6/9/12/24 months. Rs 3,000 is a ceiling, not a guaranteed amount.",
     "", "TRUE", "https://www.asus.com/in", AMEX, "2026-09-30"),
    ("Dyson", "American Express", "Electronics", "EMI", "Flat", 3000, "", "",
     "Up to Rs 3,000 instant discount on Dyson with American Express EMI.",
     "In-store 6-month EMI. Rs 3,000 is a ceiling, not a guaranteed amount.",
     "", "TRUE", "https://www.dyson.in", AMEX, "2026-08-31"),

    # --- OneCard ----------------------------------------------------------
    ("Zouk", "OneCard", "Fashion", "Credit Card", "Flat", 500, "", 2199,
     "Flat Rs 500 off at Zouk with OneCard.",
     "Minimum order Rs 2,199.",
     "ZBADN5", "FALSE", "https://www.zouk.co.in", ONECARD, "2027-03-31"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)


def autosize(ws, max_width=55):
    for col in ws.columns:
        letter = col[0].column_letter
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def main():
    wb = Workbook()

    ws = wb.active
    ws.title = "Offers"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for row in OFFERS:
        ws.append(list(row))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    autosize(ws)

    readme = wb.create_sheet("README")
    notes = [
        ["OfferLens - Batch 2 bulk upload"],
        [""],
        ["Built 30 July 2026. %d offers across 9 issuers." % len(OFFERS)],
        [""],
        ["VERIFY BEFORE PUBLISHING"],
        ["Open the 'Offer Source' URL on each row and confirm the numbers still match."],
        ["Issuers change these pages without notice and several run monthly cycles."],
        [""],
        ["HOW SOURCES WERE VERIFIED"],
        ["Read over HTTPS from the issuer's own site. Every .bank.in domain was confirmed"],
        ["by 301 redirect from the bank's long-established primary domain, not from a"],
        ["search result. No aggregator was used as a source."],
        [""],
        ["BLANK CELLS ARE DELIBERATE"],
        ["A blank Max Discount, Min Order or Valid Until means the issuer published no"],
        ["such value. These are not missing data to be filled in with an estimate."],
        [""],
        ["KNOWN LIMITS"],
        ["Standard Chartered DigiSmart offers carry no published end date - Valid Until"],
        ["is intentionally blank so they do not auto-expire."],
        ["American Express and Kotak rows are EMI offers. 'Up to Rs X' is a ceiling that"],
        ["depends on the product and tenure, not a guaranteed discount."],
        ["Bank of Baroda travel codes differ for domestic vs international; the domestic"],
        ["code is in the Coupon Code column and both are named in Terms."],
        [""],
        ["EXCLUDED ON PURPOSE"],
        ["IDFC First - every public offer expires 31 Jul 2026."],
        ["Amazon Pay - all three offers expire 1 Aug 2026 (monthly cycle)."],
        ["Paytm - offer hub is unmaintained and redirects to the homepage in a browser."],
        ["Mobikwik - expired offer pages remain publicly live; no dates published."],
        ["OneCard Bewakoof and Giva - the issuer's own feed contradicts itself on value."],
        ["Citibank - has exited Indian consumer banking; no such card exists."],
    ]
    for line in notes:
        readme.append(line)
    for row in readme.iter_rows():
        for cell in row:
            cell.font = BODY_FONT
    readme["A1"].font = Font(name="Arial", bold=True, size=13)
    readme["A5"].font = Font(name="Arial", bold=True, size=11, color="C00000")
    readme.column_dimensions["A"].width = 80

    out = "offerlens_batch2_offers.xlsx"
    wb.save(out)
    print("Wrote %s with %d offers" % (out, len(OFFERS)))


if __name__ == "__main__":
    main()
